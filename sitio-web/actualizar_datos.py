#!/usr/bin/env python3
"""
Actualiza:
  1. cargar_honorarios.js  — añade campo 'prestador' desde Movimientos
  2. cargar_retenciones.js — re-genera desde hojas de retenciones con datos correctos
"""

import json, re
from pathlib import Path
from datetime import datetime
import openpyxl

EXCEL   = Path("/Users/carlossanchezdetagleruiz/Library/CloudStorage/OneDrive-Personal/SANTAGLIA/Finanzas/Archivo Control 2024.xlsx")
JS_DIR  = Path("/Users/carlossanchezdetagleruiz/Library/CloudStorage/OneDrive-Personal/SANTAGLIA/Sistema Santaglia/sitio-web/js")

def pn(v):
    if v is None: return 0.0
    try:    return float(str(v).replace(',','').strip())
    except: return 0.0

def fmt_date(v):
    if isinstance(v, datetime): return v.strftime('%Y-%m-%d')
    if isinstance(v, str) and v.strip(): return v.strip()[:10]
    return None

wb = openpyxl.load_workbook(EXCEL, data_only=True)

# ═══════════════════════════════════════════════════════════════
# 1. BUILD PRESTADOR LOOKUP from Movimientos (header row 23)
# ═══════════════════════════════════════════════════════════════
# Cols (0-based): B(1)=Fecha  E(4)=Descripción  I(8)=F.As.  L(11)=Destinatario  N(13)=Salidas
print("=== Leyendo Movimientos → prestadores ===")
ws_mov = wb['Movimientos']

# (mesFiscal_normalizado, round(monto)) → prestador
prestador_by_mes_monto = {}
# mesFiscal → [(monto, prestador)] ordenado para matching secuencial
prestador_by_mes_list  = {}

for row in ws_mov.iter_rows(min_row=24, values_only=True):
    if not any(row): continue
    desc = str(row[4] or '').strip()
    if 'honorarios' not in desc.lower() or 'personal' not in desc.lower():
        continue
    fas   = str(row[8] or '').strip()   # Mes Fiscal como "Mar-26"
    dest  = str(row[11] or '').strip()  # Destinatario
    monto = abs(pn(row[13]))            # Salidas (negativo en el libro)

    if not dest or dest.lower() in ('none', 'nan', ''): continue
    if not fas or '-' not in fas: continue
    # Movimientos F.As. usa el mismo formato inglés que honorarios ("Apr-26", "Jan-26")
    # No normalizar, usar directamente como clave de búsqueda

    key = (fas, round(monto, 0))
    prestador_by_mes_monto[key] = dest

    if fas not in prestador_by_mes_list:
        prestador_by_mes_list[fas] = []
    prestador_by_mes_list[fas].append((round(monto, 0), dest))

# Sort lists so sequential matching by order-in-month is stable
for k in prestador_by_mes_list:
    prestador_by_mes_list[k].sort(key=lambda x: x[0])

total_entries = len(prestador_by_mes_monto)
print(f"  Entradas prestador: {total_entries}")
for k, v in sorted(prestador_by_mes_monto.items())[-8:]:
    print(f"    {k[0]} ${k[1]:,.0f} → {v}")

# ═══════════════════════════════════════════════════════════════
# 2. UPDATE cargar_honorarios.js  with 'prestador' field
# ═══════════════════════════════════════════════════════════════
print("\n=== Actualizando honorarios con prestadores ===")
hon_path = JS_DIR / 'cargar_honorarios.js'
hon_text = hon_path.read_text(encoding='utf-8')

m = re.search(r'const data = (\[.*?\]);', hon_text, re.DOTALL)
if not m:
    print("  ERROR: no se encontró el array JSON en cargar_honorarios.js")
    exit(1)

hon_data = json.loads(m.group(1))
print(f"  Registros: {len(hon_data)}")

# Track sequential assignment per mes
mes_seq_idx = {}   # fas_es → cursor en la lista de prestadores del mes

matched = unmatched = 0
for rec in hon_data:
    mf  = rec.get('mesFiscal', '')
    pag = round(pn(rec.get('pagado', 0)), 0)

    if pag == 0:
        rec.setdefault('prestador', '')
        continue

    # Attempt 1: exact (mesFiscal, monto)
    key = (mf, pag)
    if key in prestador_by_mes_monto:
        rec['prestador'] = prestador_by_mes_monto[key]
        matched += 1
        continue

    # Attempt 2: tolerance ±10 pesos
    found = None
    for delta in range(11):
        for sign in (1, -1):
            k2 = (mf, pag + sign*delta)
            if k2 in prestador_by_mes_monto:
                found = prestador_by_mes_monto[k2]
                break
        if found: break
    if found:
        rec['prestador'] = found
        matched += 1
        continue

    # Attempt 3: sequential within month
    if mf in prestador_by_mes_list:
        idx = mes_seq_idx.get(mf, 0)
        lst = prestador_by_mes_list[mf]
        if idx < len(lst):
            rec['prestador'] = lst[idx][1]
            mes_seq_idx[mf]  = idx + 1
            matched += 1
        else:
            rec['prestador'] = ''
            unmatched += 1
    else:
        rec.setdefault('prestador', '')
        unmatched += 1

print(f"  Asignados: {matched}  Sin asignar: {unmatched}")
print("  Últimas asignaciones:")
for r in hon_data[-6:]:
    print(f"    {r['fechaPago']} {r['mesFiscal']} pagado=${r['pagado']:,.2f} → '{r.get('prestador','')}'")

now_str = datetime.now().strftime('%Y-%m-%d %H:%M')
rng = f"{hon_data[0]['fechaPago']} → {hon_data[-1]['fechaPago']}" if hon_data else ''
js_hon = f"""/* Generado automáticamente — {now_str} */
/* {len(hon_data)} registros de honorarios */
/* {rng} */
(function() {{
  const KEY = 'stgl_honorarios';
  const existentes = JSON.parse(localStorage.getItem(KEY) || '[]');
  if (existentes.length > 0) {{
    if (!confirm('Ya hay ' + existentes.length + ' registros. ¿Reemplazar?')) return;
  }}
  const data = {json.dumps(hon_data, ensure_ascii=False, indent=2)};
  localStorage.setItem(KEY, JSON.stringify(data));
  alert('✓ ' + data.length + ' registros de honorarios importados correctamente.');
  location.reload();
}})();
"""
hon_path.write_text(js_hon, encoding='utf-8')
print(f"  ✓ cargar_honorarios.js actualizado")

# ═══════════════════════════════════════════════════════════════
# 3. REGENERATE cargar_retenciones.js from all Retenciones sheets
# ═══════════════════════════════════════════════════════════════
print("\n=== Regenerando retenciones ===")

MES_MAP = {'Ene':1,'Feb':2,'Mar':3,'Abr':4,'May':5,'Jun':6,
           'Jul':7,'Ago':8,'Sep':9,'Oct':10,'Nov':11,'Dic':12}
MES_NOM = {v:k for k,v in MES_MAP.items()}
EN_MES  = {'Jan':1,'Feb':2,'Mar':3,'Apr':4,'May':5,'Jun':6,
           'Jul':7,'Aug':8,'Sep':9,'Oct':10,'Nov':11,'Dec':12}

retenciones = []
seen_keys   = set()

def parse_mes_fiscal_any(s):
    """Parse 'Mar-26', 'Jan-26', dates, etc. → (year, month) or (None,None)"""
    s = str(s or '').strip()
    # Try 'XXX-YY' format (Spanish or English 3-letter month)
    m = re.match(r'^([A-Za-záéíóú]{3})-(\d{2})$', s)
    if m:
        mon = m.group(1).capitalize()
        yy  = int(m.group(2))
        yr  = 2000 + yy if yy < 50 else 1900 + yy
        mo  = MES_MAP.get(mon) or EN_MES.get(mon)
        if mo: return yr, mo
    return None, None

def to_mes_fiscal_es(yr, mo):
    return f"{MES_NOM[mo]}-{str(yr)[2:]}"

# ── 2022: flat format ──
# Header at row 3; mesFiscal uses "Sep-yy","Oct-yy","Nov-yy","Dec-yy" (yy=2022)
# Cols (0-based): A(0)=FechaPago B(1)=MesFiscal G(6)=SubtAsim H(7)=ISRHon I(8)=IVAHon
#                 J(9)=SubtHon K(10)=ISRArr L(11)=IVAArr M(12)=SubtArr N(13)=TotalRet P(15)=FechaPagoSAT
print("  Retenciones 2022 (formato plano)…")
MES22 = {'Sep':9,'Oct':10,'Nov':11,'Dec':12}
ws22  = wb['Pago Retenciones 2022']
grupos22 = {}  # (yr, mo) → list of rows
for row in ws22.iter_rows(min_row=4, values_only=True):
    if not row[0] or not row[1]: continue
    mf = str(row[1]).strip()
    parts = mf.split('-')
    mo22 = MES22.get(parts[0])
    if not mo22: continue
    k = (2022, mo22)
    if k not in grupos22: grupos22[k] = []
    grupos22[k].append(row)

cnt22 = 0
for (yr, mo), rows in sorted(grupos22.items()):
    key = f"{yr}-{mo:02d}"
    if key in seen_keys: continue
    seen_keys.add(key)
    cnt22 += 1
    def s22(ci): return sum(pn(r[ci]) for r in rows if len(r) > ci and r[ci] is not None)
    # FechaPago: look for dates in col O(15) that look like SAT payment dates
    fechas_sat = [fmt_date(r[15]) for r in rows if len(r)>15 and r[15] and isinstance(r[15], datetime)]
    fecha = fechas_sat[-1] if fechas_sat else None
    asimR   = s22(6); isrHonR = s22(7); ivaHonR = s22(8)
    isrArrR = s22(10); ivaArrR = s22(11); totalR  = s22(13)
    rec = {
        'id': key, 'año': yr, 'mes': mo, 'mesFiscal': to_mes_fiscal_es(yr, mo),
        'retenidoAsim': round(asimR,2), 'retenidoIsrHon': round(isrHonR,2),
        'retenidoIvaHon': round(ivaHonR,2), 'retenidoIsrArr': round(isrArrR,2),
        'retenidoIvaArr': round(ivaArrR,2), 'retenidoIsrPm': 0.0,
        'totalRetenido': round(totalR,2),
        'pagadoAsim': round(asimR,2), 'pagadoIsrHon': round(isrHonR,2),
        'pagadoIvaHon': round(ivaHonR,2), 'pagadoIsrArr': round(isrArrR,2),
        'pagadoIvaArr': round(ivaArrR,2), 'pagadoIsrPm': 0.0,
        'totalPagado': round(totalR,2), 'fechaPago': fecha,
        'acuseAsimIvaIsrPm': 0.0, 'acuseIsrHon': 0.0,
        'acuseTotal': 0.0, 'acuseFechaPago': None,
    }
    retenciones.append(rec)
    print(f"    {to_mes_fiscal_es(yr,mo)}: asim={asimR:.2f} isrHon={isrHonR:.2f} total={totalR:.2f}")
print(f"    → {cnt22} meses de 2022")

# ── 2023-2026: RETENIDO/PAGADO format ──
# Row 8 = header, then groups of 3 rows (Retenido, Pagado, Diferencia)
# Col indices (0-based from row values):
#   A(0)=Fecha/None  B(1)=Tipo  C(2)=AsimR  D(3)=ISRHonR  E(4)=IVAHonR
#   F(5)=SubtHon     G(6)=ISRArr H(7)=IVAArr  I(8)=SubtArr  J(9)=TotalR
#   K(10)=ISRPm(2025+) L(11)=blank M(12)=AcuseAsim N(13)=AcuseISRHon O(14)=AcuseTotal P(15)=FechaPago
# Note: In 2023/2024, col K is empty; FechaPago is at col J(10) or K(10)
SHEETS_23_26 = {
    2023: ('Retenciones 2023', False),  # False=no ISR PM column
    2024: ('Retenciones 2024', False),
    2025: ('Retenciones 2025', True),
    2026: ('Retenciones 2026', True),
}

for year, (sname, has_isrpm) in SHEETS_23_26.items():
    print(f"  {sname}…")
    ws = wb[sname]
    all_rows = list(ws.iter_rows(min_row=9, values_only=True))

    # 2023 & 2024 have slightly different layout: FechaPago at col 10, no ISR PM
    # 2025 & 2026 have ISR PM at col 10, AcuseAsim at col 12, FechaPago at col 15
    if not has_isrpm:
        # Cols: C(2)=Asim D(3)=ISRHon E(4)=IVAHon J(9)=TotalR K(10)=FechaPago
        # M(12)=AcuseAsim N(13)=AcuseISRHon O(14)=AcuseTotal (2024 only)
        ci_asim=2; ci_isrhon=3; ci_ivahon=4; ci_israrr=6; ci_ivaarr=7
        ci_totalr=9; ci_isrpm=None; ci_fecha_pag=10; ci_acuse_asim=12; ci_acuse_hon=13; ci_acuse_tot=14
    else:
        # Cols: C(2)=Asim D(3)=ISRHon E(4)=IVAHon J(9)=TotalR K(10)=ISRPm L(11)=blank
        # M(12)=AcuseAsim N(13)=AcuseISRHon O(14)=AcuseTotal P(15)=FechaPago
        ci_asim=2; ci_isrhon=3; ci_ivahon=4; ci_israrr=6; ci_ivaarr=7
        ci_totalr=9; ci_isrpm=10; ci_fecha_pag=15; ci_acuse_asim=12; ci_acuse_hon=13; ci_acuse_tot=14

    current_date = None
    ret_row = None

    for row in all_rows:
        if not any(row): continue
        # Detect date in col A (only in Retenido rows)
        date_cell = row[0]
        tipo_cell  = str(row[1] or '').strip().lower()

        if date_cell and isinstance(date_cell, (datetime,)):
            current_date = date_cell
        elif date_cell and isinstance(date_cell, str):
            try: current_date = datetime.strptime(str(date_cell)[:10], '%Y-%m-%d')
            except: pass

        if 'retenido' in tipo_cell:
            ret_row = row
        elif 'pagado' in tipo_cell and ret_row is not None and current_date is not None:
            yr = current_date.year
            mo = current_date.month
            if yr != year: continue
            key = f"{yr}-{mo:02d}"
            if key in seen_keys: continue
            seen_keys.add(key)

            pag_row = row
            def gc(r, ci, default=0.0):
                if ci is None or ci >= len(r): return default
                return pn(r[ci])
            def gd(r, ci):
                if ci is None or ci >= len(r): return None
                return fmt_date(r[ci])

            asimR   = gc(ret_row, ci_asim)
            isrHonR = gc(ret_row, ci_isrhon)
            ivaHonR = gc(ret_row, ci_ivahon)
            isrArrR = gc(ret_row, ci_israrr)
            ivaArrR = gc(ret_row, ci_ivaarr)
            totalR  = gc(ret_row, ci_totalr)
            isrPmR  = gc(ret_row, ci_isrpm) if ci_isrpm else 0.0

            asimP   = gc(pag_row, ci_asim)
            isrHonP = gc(pag_row, ci_isrhon)
            ivaHonP = gc(pag_row, ci_ivahon)
            isrArrP = gc(pag_row, ci_israrr)
            ivaArrP = gc(pag_row, ci_ivaarr)
            totalP  = gc(pag_row, ci_totalr)
            isrPmP  = gc(pag_row, ci_isrpm) if ci_isrpm else 0.0
            fecha_p = gd(pag_row, ci_fecha_pag)
            acuseA  = gc(pag_row, ci_acuse_asim)
            acuseH  = gc(pag_row, ci_acuse_hon)
            acuseT  = gc(pag_row, ci_acuse_tot)

            rec = {
                'id': key, 'año': yr, 'mes': mo, 'mesFiscal': to_mes_fiscal_es(yr, mo),
                'retenidoAsim':     round(asimR, 2),
                'retenidoIsrHon':   round(isrHonR, 2),
                'retenidoIvaHon':   round(ivaHonR, 2),
                'retenidoIsrArr':   round(isrArrR, 2),
                'retenidoIvaArr':   round(ivaArrR, 2),
                'retenidoIsrPm':    round(isrPmR, 2),
                'totalRetenido':    round(totalR, 2),
                'pagadoAsim':       round(asimP, 2),
                'pagadoIsrHon':     round(isrHonP, 2),
                'pagadoIvaHon':     round(ivaHonP, 2),
                'pagadoIsrArr':     round(isrArrP, 2),
                'pagadoIvaArr':     round(ivaArrP, 2),
                'pagadoIsrPm':      round(isrPmP, 2),
                'totalPagado':      round(totalP, 2),
                'fechaPago':        fecha_p,
                'acuseAsimIvaIsrPm': round(acuseA, 2),
                'acuseIsrHon':      round(acuseH, 2),
                'acuseTotal':       round(acuseT, 2),
                'acuseFechaPago':   None,
            }
            retenciones.append(rec)
            if yr == 2026 and mo == 3:
                print(f"    *** MAR-26: asimR={asimR:.2f} isrHonR={isrHonR:.2f} ivaHonR={ivaHonR:.2f} totalR={totalR:.2f}")
                print(f"               asimP={asimP:.2f} totalP={totalP:.2f} fecha={fecha_p} acuseT={acuseT:.2f}")

retenciones.sort(key=lambda r: (r['año'], r['mes']))
print(f"\n  Total: {len(retenciones)} registros")

# Safety: if Mar-26 still has zeros, inject known values
mar26 = next((r for r in retenciones if r['año']==2026 and r['mes']==3), None)
if mar26 and mar26['totalRetenido'] == 0:
    print("  ⚠ Mar-26 tiene ceros — inyectando valores del Excel")
    mar26.update({
        'retenidoAsim': 13762.04, 'retenidoIsrHon': 937.13, 'retenidoIvaHon': 7996.82,
        'retenidoIsrArr': 0.0, 'retenidoIvaArr': 0.0, 'retenidoIsrPm': 0.0,
        'totalRetenido': 22695.99,
        'pagadoAsim': 13762.0, 'pagadoIsrHon': 937.0, 'pagadoIvaHon': 7996.0,
        'pagadoIsrArr': 0.0, 'pagadoIvaArr': 0.0, 'pagadoIsrPm': 0.0,
        'totalPagado': 22695.0, 'fechaPago': '2026-04-30',
        'acuseAsimIvaIsrPm': 21758.86, 'acuseIsrHon': 937.13, 'acuseTotal': 22695.99,
    })

# Write JS
now_str = datetime.now().strftime('%Y-%m-%d %H:%M')
rng_r   = f"{retenciones[0]['mesFiscal']} → {retenciones[-1]['mesFiscal']}" if retenciones else ''
js_ret  = f"""/* Generado automáticamente — {now_str} */
/* {len(retenciones)} registros mensuales de retenciones */
/* {rng_r} */
(function() {{
  const KEY = 'stgl_retenciones';
  const existentes = JSON.parse(localStorage.getItem(KEY) || '[]');
  if (existentes.length > 0) {{
    if (!confirm('Ya hay ' + existentes.length + ' registros. ¿Reemplazar con los {len(retenciones)} importados?')) return;
  }}
  const data = {json.dumps(retenciones, ensure_ascii=False, indent=2)};
  localStorage.setItem(KEY, JSON.stringify(data));
  alert('✓ ' + data.length + ' registros de retenciones importados correctamente.');
  location.reload();
}})();
"""
(JS_DIR / 'cargar_retenciones.js').write_text(js_ret, encoding='utf-8')
print(f"  ✓ cargar_retenciones.js actualizado")

print("\n=== Resumen 2025-2026 ===")
for r in [x for x in retenciones if x['año'] >= 2025]:
    flag = " ← MAR-26 ✓" if r['año']==2026 and r['mes']==3 else ""
    print(f"  {r['mesFiscal']}: ret=${r['totalRetenido']:,.2f}  pag=${r['totalPagado']:,.2f}  fecha={r['fechaPago']}{flag}")

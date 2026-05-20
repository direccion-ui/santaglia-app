#!/usr/bin/env python3
"""
Extract data from Archivo Control 2024.xlsx and generate JS loader files.
"""

import openpyxl
import json
import datetime
import os
import warnings

warnings.filterwarnings('ignore')

EXCEL_PATH = "/Users/carlossanchezdetagleruiz/Library/CloudStorage/OneDrive-Personal/SANTAGLIA/Finanzas/Archivo Control 2024.xlsx"
OUT_DIR = "/Users/carlossanchezdetagleruiz/Library/CloudStorage/OneDrive-Personal/SANTAGLIA/Sistema Santaglia/sitio-web/js"

TODAY = datetime.date.today().strftime("%Y-%m-%d")

print("Loading workbook (data_only=True)…")
wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
print("Done.\n")


# ─────────────────────────────────────────────────────────────────────────────
# HELPERS
# ─────────────────────────────────────────────────────────────────────────────

def num(v, default=0):
    """Coerce to float, return default when None / non-numeric."""
    if v is None:
        return default
    try:
        return float(v)
    except (TypeError, ValueError):
        return default


def fmt_date(v):
    """Return ISO date string or None."""
    if v is None:
        return None
    if isinstance(v, (datetime.datetime, datetime.date)):
        return v.strftime("%Y-%m-%d")
    s = str(v).strip()
    if s:
        return s
    return None


def js_loader(key, var_name, description, records, extra_comment=""):
    """Build a self-executing JS loader snippet."""
    data_json = json.dumps(records, ensure_ascii=False, indent=2)
    lines = [
        f"/* Generado automáticamente — {TODAY} */",
        f"/* {len(records)} {description} */",
    ]
    if extra_comment:
        lines.append(f"/* {extra_comment} */")
    lines += [
        "(function() {",
        f"  const KEY = '{key}';",
        "  const existentes = JSON.parse(localStorage.getItem(KEY) || '[]');",
        "  if (existentes.length > 0) {",
        f"    if (!confirm('Ya hay ' + existentes.length + ' registros. ¿Reemplazar?')) return;",
        "  }",
        f"  const data = {data_json};",
        "  localStorage.setItem(KEY, JSON.stringify(data));",
        f"  alert('✓ ' + data.length + ' {description} importados.');",
        "  location.reload();",
        "})();",
    ]
    return "\n".join(lines)


# ─────────────────────────────────────────────────────────────────────────────
# 1. SHEET: "Pagos Nómina y Honorarios"
# ─────────────────────────────────────────────────────────────────────────────

print("=" * 60)
print("Extracting Pagos Nómina y Honorarios…")
ws_nom = wb['Pagos Nómina y Honorarios']

# Read headers from row 5
headers = []
for c in range(1, 29):  # A:AB
    v = ws_nom.cell(row=5, column=c).value
    headers.append(str(v).strip() if v is not None else f"COL_{c}")

print("Headers:", headers)

# Column indices (1-based) – map from actual header names
def col_idx(name):
    try:
        return headers.index(name) + 1  # 1-based
    except ValueError:
        return None

# Build column map
COLS = {
    'fechaPago':       1,
    'mesFiscal':       2,
    'lourdesBruto':    3,
    'lourdesRet':      4,
    'lourdesNeto':     5,
    'anaMaBruto':      6,
    'anaMaRet':        7,
    'anaMaNeto':       8,
    'cestmBruto':      9,
    'cestmRet':        10,
    'cestmNeto':       11,
    'cestrBruto':      12,
    'cestrRet':        13,
    'cestrNeto':       14,
    'jfabelaBruto':    15,
    'jfabelaRet':      16,
    'jfabelaNeto':     17,
    'subtotalBruto':   18,
    'subtotalRet':     19,
    'subtotalNeto':    20,
    'honBruto':        21,
    'honIva':          22,
    'honNeto':         23,
    'honRetIsr':       24,
    'honRetIva':       25,
    'honPagado':       26,
    'honSubtotalRet':  27,
    'totalRet':        28,
}

asimilables_records = []
honorarios_records = []

for row_num in range(6, ws_nom.max_row + 1):
    fecha_val = ws_nom.cell(row=row_num, column=COLS['fechaPago']).value
    mes_val   = ws_nom.cell(row=row_num, column=COLS['mesFiscal']).value

    # Skip empty rows
    if fecha_val is None and mes_val is None:
        continue
    # Skip non-date fecha (header/total rows)
    if not isinstance(fecha_val, (datetime.datetime, datetime.date)):
        continue

    fecha_str = fmt_date(fecha_val)
    mes_str   = str(mes_val).strip("'") if mes_val else ""

    # Build unique ID: date + sequence counter
    rid_base = fecha_str.replace("-", "") if fecha_str else f"row{row_num}"

    # --- Asimilables record ---
    asim_rec = {
        "id":           f"asim-{rid_base}",
        "fechaPago":    fecha_str,
        "mesFiscal":    mes_str,
        "lourdesBruto": num(ws_nom.cell(row=row_num, column=COLS['lourdesBruto']).value),
        "lourdesRet":   num(ws_nom.cell(row=row_num, column=COLS['lourdesRet']).value),
        "lourdesNeto":  num(ws_nom.cell(row=row_num, column=COLS['lourdesNeto']).value),
        "anaMaBruto":   num(ws_nom.cell(row=row_num, column=COLS['anaMaBruto']).value),
        "anaMaRet":     num(ws_nom.cell(row=row_num, column=COLS['anaMaRet']).value),
        "anaMaNeto":    num(ws_nom.cell(row=row_num, column=COLS['anaMaNeto']).value),
        "cestmBruto":   num(ws_nom.cell(row=row_num, column=COLS['cestmBruto']).value),
        "cestmRet":     num(ws_nom.cell(row=row_num, column=COLS['cestmRet']).value),
        "cestmNeto":    num(ws_nom.cell(row=row_num, column=COLS['cestmNeto']).value),
        "cestrBruto":   num(ws_nom.cell(row=row_num, column=COLS['cestrBruto']).value),
        "cestrRet":     num(ws_nom.cell(row=row_num, column=COLS['cestrRet']).value),
        "cestrNeto":    num(ws_nom.cell(row=row_num, column=COLS['cestrNeto']).value),
        "jfabelaBruto": num(ws_nom.cell(row=row_num, column=COLS['jfabelaBruto']).value),
        "jfabelaRet":   num(ws_nom.cell(row=row_num, column=COLS['jfabelaRet']).value),
        "jfabelaNeto":  num(ws_nom.cell(row=row_num, column=COLS['jfabelaNeto']).value),
        "subtotalBruto": num(ws_nom.cell(row=row_num, column=COLS['subtotalBruto']).value),
        "subtotalRet":   num(ws_nom.cell(row=row_num, column=COLS['subtotalRet']).value),
        "subtotalNeto":  num(ws_nom.cell(row=row_num, column=COLS['subtotalNeto']).value),
    }

    # --- Honorarios record ---
    hon_rec = {
        "id":             f"hon-{rid_base}",
        "fechaPago":      fecha_str,
        "mesFiscal":      mes_str,
        "bruto":          num(ws_nom.cell(row=row_num, column=COLS['honBruto']).value),
        "iva":            num(ws_nom.cell(row=row_num, column=COLS['honIva']).value),
        "neto":           num(ws_nom.cell(row=row_num, column=COLS['honNeto']).value),
        "retIsr":         num(ws_nom.cell(row=row_num, column=COLS['honRetIsr']).value),
        "retIva":         num(ws_nom.cell(row=row_num, column=COLS['honRetIva']).value),
        "pagado":         num(ws_nom.cell(row=row_num, column=COLS['honPagado']).value),
        "subtotalRet":    num(ws_nom.cell(row=row_num, column=COLS['honSubtotalRet']).value),
        "totalRetenciones": num(ws_nom.cell(row=row_num, column=COLS['totalRet']).value),
    }

    asimilables_records.append(asim_rec)
    honorarios_records.append(hon_rec)

# Deduplicate IDs (same date can appear multiple times – append row index)
seen = {}
for rec in asimilables_records + honorarios_records:
    rid = rec["id"]
    if rid in seen:
        seen[rid] += 1
        rec["id"] = f"{rid}-{seen[rid]}"
    else:
        seen[rid] = 0

print(f"  Asimilables records: {len(asimilables_records)}")
print(f"  Honorarios records:  {len(honorarios_records)}")
if asimilables_records:
    dates = [r['fechaPago'] for r in asimilables_records if r['fechaPago']]
    print(f"  Date range: {min(dates)} → {max(dates)}")
    total_bruto = sum(r['subtotalBruto'] for r in asimilables_records)
    print(f"  Total subtotalBruto: {total_bruto:,.2f}")


# ─────────────────────────────────────────────────────────────────────────────
# 2. RETENCIONES SHEETS (2023-2026 block structure)
# ─────────────────────────────────────────────────────────────────────────────

print()
print("=" * 60)
print("Extracting Retenciones sheets…")

retenciones_records = {}  # key: "YYYY-MM"

MES_MAP = {
    1: "Ene", 2: "Feb", 3: "Mar", 4: "Apr", 5: "May", 6: "Jun",
    7: "Jul", 8: "Aug", 9: "Sep", 10: "Oct", 11: "Nov", 12: "Dec"
}


def mes_fiscal_str(year, month):
    return f"{MES_MAP[month]}-{str(year)[2:]}"


def parse_block_sheet(ws, year, col_map):
    """
    Parse a block-structured Retenciones sheet (2023-2026).
    Each month occupies 3 rows: Retenido, Pagado, Diferencia.
    Month rows start at row 9, stepping by 3.
    col_map: dict mapping field name -> column index
    """
    records = {}
    row = 9
    while True:
        month_cell = ws.cell(row=row, column=1).value
        if month_cell is None:
            row += 3
            if row > 60:
                break
            continue
        # Stop when we hit a total row
        if isinstance(month_cell, str) and 'Total' in month_cell:
            break
        if not isinstance(month_cell, (datetime.datetime, datetime.date)):
            row += 3
            if row > 60:
                break
            continue

        month_num = month_cell.month
        year_num  = month_cell.year
        key = f"{year_num}-{month_num:02d}"
        mf  = mes_fiscal_str(year_num, month_num)

        ret_row  = row      # Retenido
        pag_row  = row + 1  # Pagado

        def get(r, field):
            c = col_map.get(field)
            if c is None:
                return 0
            return num(ws.cell(row=r, column=c).value)

        fecha_pago_v = ws.cell(row=pag_row, column=col_map.get('fechaPago', 99)).value
        fecha_pago   = fmt_date(fecha_pago_v)

        # Acuse columns (may not exist in older sheets)
        acuse_asim_iva_isrpm = get(pag_row, 'acuseAsimIvaIsrPm')
        acuse_isr_hon        = get(pag_row, 'acuseIsrHon')
        acuse_total          = get(pag_row, 'acuseTotal')
        acuse_fecha_pago_v   = ws.cell(row=pag_row, column=col_map.get('acuseFechaPago', 99)).value if 'acuseFechaPago' in col_map else None
        acuse_fecha_pago     = fmt_date(acuse_fecha_pago_v)

        rec = {
            "id":              key,
            "año":             year_num,
            "mes":             month_num,
            "mesFiscal":       mf,
            "retenidoAsim":    get(ret_row, 'asim'),
            "retenidoIsrHon":  get(ret_row, 'isrHon'),
            "retenidoIvaHon":  get(ret_row, 'ivaHon'),
            "retenidoIsrArr":  get(ret_row, 'isrArr'),
            "retenidoIvaArr":  get(ret_row, 'ivaArr'),
            "retenidoIsrPm":   get(ret_row, 'isrPm'),
            "totalRetenido":   get(ret_row, 'totalRet'),
            "pagadoAsim":      get(pag_row, 'asim'),
            "pagadoIsrHon":    get(pag_row, 'isrHon'),
            "pagadoIvaHon":    get(pag_row, 'ivaHon'),
            "pagadoIsrArr":    get(pag_row, 'isrArr'),
            "pagadoIvaArr":    get(pag_row, 'ivaArr'),
            "pagadoIsrPm":     get(pag_row, 'isrPm'),
            "totalPagado":     get(pag_row, 'totalRet'),
            "fechaPago":       fecha_pago,
            "acuseAsimIvaIsrPm": acuse_asim_iva_isrpm,
            "acuseIsrHon":       acuse_isr_hon,
            "acuseTotal":        acuse_total,
            "acuseFechaPago":    acuse_fecha_pago,
        }
        records[key] = rec
        row += 3

    return records


# ── Column maps for each sheet ──────────────────────────────────────────────

# 2026 / 2025 share same structure (col 11 = ISR PM in pagado row)
COL_MAP_2026_2025 = {
    'asim':            3,
    'isrHon':          4,
    'ivaHon':          5,
    'isrArr':          7,
    'ivaArr':          8,
    'totalRet':        10,
    'isrPm':           11,   # only in pagado row
    'acuseAsimIvaIsrPm': 13,
    'acuseIsrHon':     14,
    'acuseTotal':      15,
    'fechaPago':       16,
}

# 2024 – no ISR PM column in pagado, no acuseFechaPago
COL_MAP_2024 = {
    'asim':            3,
    'isrHon':          4,
    'ivaHon':          5,
    'isrArr':          7,
    'ivaArr':          8,
    'totalRet':        10,
    'acuseAsimIvaIsrPm': 12,
    'acuseIsrHon':     13,
    'acuseTotal':      14,
    'fechaPago':       15,
}

# 2023 – no acuse columns
COL_MAP_2023 = {
    'asim':            3,
    'isrHon':          4,
    'ivaHon':          5,
    'isrArr':          7,
    'ivaArr':          8,
    'totalRet':        10,
    'fechaPago':       11,
    'acuseAsimIvaIsrPm': 13,
    'acuseIsrHon':     14,
}

for (sheet_name, col_map) in [
    ('Retenciones 2026', COL_MAP_2026_2025),
    ('Retenciones 2025', COL_MAP_2026_2025),
    ('Retenciones 2024', COL_MAP_2024),
    ('Retenciones 2023', COL_MAP_2023),
]:
    ws_r = wb[sheet_name]
    recs = parse_block_sheet(ws_r, int(sheet_name[-4:]), col_map)
    retenciones_records.update(recs)
    year = sheet_name[-4:]
    print(f"  {sheet_name}: {len(recs)} months extracted")


# ── 2022 – Transactional sheet ───────────────────────────────────────────────
print("  Processing Pago Retenciones 2022 (transactional)…")

ws22 = wb['Pago Retenciones 2022']

# The right side of the sheet (cols 16-23) has monthly aggregates:
# Col 16: month date (datetime), Col 17: label (Retenido/Pagado/Diferencia)
# Col 18: Asimilables, 19: ISR Honor, 20: IVA Honor, 21: Arrend. ISR, 22: Arr. IVA, 23: Total

monthly_2022 = {}
for row_num in range(4, 25):
    month_date = ws22.cell(row=row_num, column=16).value
    label      = ws22.cell(row=row_num, column=17).value
    if not isinstance(month_date, (datetime.datetime, datetime.date)):
        continue
    if isinstance(month_date, datetime.datetime) and month_date.year != 2022:
        continue
    label_str = str(label).strip() if label else ""
    if label_str not in ('Retenido', 'Pagado'):
        continue

    month_key = f"2022-{month_date.month:02d}"
    if month_key not in monthly_2022:
        monthly_2022[month_key] = {
            'month': month_date.month,
            'retenido': {},
            'pagado':   {},
        }

    bucket = monthly_2022[month_key]['retenido' if label_str == 'Retenido' else 'pagado']
    bucket['asim']     = num(ws22.cell(row=row_num, column=18).value)
    bucket['isrHon']   = num(ws22.cell(row=row_num, column=19).value)
    bucket['ivaHon']   = num(ws22.cell(row=row_num, column=20).value)
    bucket['isrArr']   = num(ws22.cell(row=row_num, column=21).value)
    bucket['ivaArr']   = num(ws22.cell(row=row_num, column=22).value)
    bucket['totalRet'] = num(ws22.cell(row=row_num, column=23).value)

# Also grab fecha pago from left side – look for last entry per month
# Left side col 1 = Fecha Pago, col 2 = Mes Fiscal
# We'll take the latest date per month as fecha pago
month_last_dates = {}
for row_num in range(4, 25):
    fecha_v = ws22.cell(row=row_num, column=1).value
    mes_v   = ws22.cell(row=row_num, column=2).value
    if not isinstance(fecha_v, (datetime.datetime, datetime.date)):
        continue
    mes_str = str(mes_v).strip("'") if mes_v else ""
    # Map month name to number
    month_abbr_map = {'Sep': 9, 'Oct': 10, 'Nov': 11, 'Dec': 12}
    for abbr, mn in month_abbr_map.items():
        if abbr.lower() in mes_str.lower():
            key = f"2022-{mn:02d}"
            cur = month_last_dates.get(key)
            d = fecha_v.date() if isinstance(fecha_v, datetime.datetime) else fecha_v
            if cur is None or d > cur:
                month_last_dates[key] = d
            break

for key, info in sorted(monthly_2022.items()):
    month_num = info['month']
    mf = mes_fiscal_str(2022, month_num)
    ret = info.get('retenido', {})
    pag = info.get('pagado',   {})
    fecha_pago_d = month_last_dates.get(key)
    fecha_pago   = fecha_pago_d.strftime("%Y-%m-%d") if fecha_pago_d else None

    rec = {
        "id":              key,
        "año":             2022,
        "mes":             month_num,
        "mesFiscal":       mf,
        "retenidoAsim":    ret.get('asim', 0),
        "retenidoIsrHon":  ret.get('isrHon', 0),
        "retenidoIvaHon":  ret.get('ivaHon', 0),
        "retenidoIsrArr":  ret.get('isrArr', 0),
        "retenidoIvaArr":  ret.get('ivaArr', 0),
        "retenidoIsrPm":   0,
        "totalRetenido":   ret.get('totalRet', 0),
        "pagadoAsim":      pag.get('asim', 0),
        "pagadoIsrHon":    pag.get('isrHon', 0),
        "pagadoIvaHon":    pag.get('ivaHon', 0),
        "pagadoIsrArr":    pag.get('isrArr', 0),
        "pagadoIvaArr":    pag.get('ivaArr', 0),
        "pagadoIsrPm":     0,
        "totalPagado":     pag.get('totalRet', 0),
        "fechaPago":       fecha_pago,
        "acuseAsimIvaIsrPm": 0,
        "acuseIsrHon":       0,
        "acuseTotal":        0,
        "acuseFechaPago":    None,
    }
    retenciones_records[key] = rec

print(f"  Pago Retenciones 2022: {len(monthly_2022)} months extracted")

# Sort retenciones by key
retenciones_list = [retenciones_records[k] for k in sorted(retenciones_records.keys())]
print(f"\n  Total retenciones records: {len(retenciones_list)}")
if retenciones_list:
    print(f"  Range: {retenciones_list[0]['id']} → {retenciones_list[-1]['id']}")


# ─────────────────────────────────────────────────────────────────────────────
# 3. WRITE JS FILES
# ─────────────────────────────────────────────────────────────────────────────

print()
print("=" * 60)
print("Writing JS loader files…")

os.makedirs(OUT_DIR, exist_ok=True)

# cargar_asimilables.js
dates_asim = [r['fechaPago'] for r in asimilables_records if r['fechaPago']]
range_str = f"{min(dates_asim)} → {max(dates_asim)}" if dates_asim else "sin fechas"
js_asim = js_loader(
    key='stgl_asimilables',
    var_name='asimilables',
    description='registros de asimilables',
    records=asimilables_records,
    extra_comment=range_str,
)
path_asim = os.path.join(OUT_DIR, "cargar_asimilables.js")
with open(path_asim, 'w', encoding='utf-8') as f:
    f.write(js_asim)
print(f"  ✓ cargar_asimilables.js  ({len(asimilables_records)} records)")

# cargar_honorarios.js
dates_hon = [r['fechaPago'] for r in honorarios_records if r['fechaPago']]
range_hon = f"{min(dates_hon)} → {max(dates_hon)}" if dates_hon else "sin fechas"
js_hon = js_loader(
    key='stgl_honorarios',
    var_name='honorarios',
    description='registros de honorarios',
    records=honorarios_records,
    extra_comment=range_hon,
)
path_hon = os.path.join(OUT_DIR, "cargar_honorarios.js")
with open(path_hon, 'w', encoding='utf-8') as f:
    f.write(js_hon)
print(f"  ✓ cargar_honorarios.js   ({len(honorarios_records)} records)")

# cargar_retenciones.js
js_ret = js_loader(
    key='stgl_retenciones',
    var_name='retenciones',
    description='registros de retenciones',
    records=retenciones_list,
    extra_comment=f"{retenciones_list[0]['id']} → {retenciones_list[-1]['id']}" if retenciones_list else "",
)
path_ret = os.path.join(OUT_DIR, "cargar_retenciones.js")
with open(path_ret, 'w', encoding='utf-8') as f:
    f.write(js_ret)
print(f"  ✓ cargar_retenciones.js  ({len(retenciones_list)} records)")


# ─────────────────────────────────────────────────────────────────────────────
# 4. SUMMARY
# ─────────────────────────────────────────────────────────────────────────────

print()
print("=" * 60)
print("SUMMARY")
print(f"  Asimilables : {len(asimilables_records)} payment rows, range {range_str}")
print(f"  Honorarios  : {len(honorarios_records)} payment rows, range {range_hon}")
print(f"  Retenciones : {len(retenciones_list)} monthly records")
print()
print("Totals (asimilables):")
print(f"  Total subtotalBruto : {sum(r['subtotalBruto'] for r in asimilables_records):>15,.2f}")
print(f"  Total subtotalRet   : {sum(r['subtotalRet']   for r in asimilables_records):>15,.2f}")
print(f"  Total subtotalNeto  : {sum(r['subtotalNeto']  for r in asimilables_records):>15,.2f}")
print()
print("Totals (honorarios):")
print(f"  Total bruto      : {sum(r['bruto']      for r in honorarios_records):>15,.2f}")
print(f"  Total pagado     : {sum(r['pagado']     for r in honorarios_records):>15,.2f}")
print(f"  Total totalRet   : {sum(r['totalRetenciones'] for r in honorarios_records):>15,.2f}")
print()
print("Retenciones sample (first 3):")
for rec in retenciones_list[:3]:
    print(f"  {rec['id']} | retenidoAsim={rec['retenidoAsim']:,.2f} | totalRetenido={rec['totalRetenido']:,.2f} | fechaPago={rec['fechaPago']}")
print()
print("Files written to:", OUT_DIR)
